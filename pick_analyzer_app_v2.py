import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import io
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN FIJA
# Actualiza estas listas si hay cambios de personal
# ============================================================
COLUMNA_PICKLIST  = "Picklist Code"
COLUMNA_ASSOCIATE = "Associate"
COLUMNA_BAGS      = "Bags"
COLUMNA_OVS       = "OVs"
COLUMNA_DURATION  = "Duration"

YM_LOGINS = [
    "vflorezf", "zamarren", "zurerad", "evalinva", "elisabgo"
]

MANAGER_LOGINS = [
    "sircam", "vajul", "aalonmu", "rplazas", "velasjuz",
    "boixadep", "rodrfali", "elearnau", "penelmax", "matapaj"
]

ROLES_NO_AA = set([x.lower() for x in YM_LOGINS + MANAGER_LOGINS])


# ============================================================
# UTILIDADES
# ============================================================
def duration_a_segundos(valor):
    try:
        partes   = str(valor).split(":")
        horas    = int(partes[0])
        minutos  = int(partes[1])
        segundos = int(partes[2].split(".")[0])
        return horas * 3600 + minutos * 60 + segundos
    except Exception:
        try:
            return float(valor) * 86400
        except Exception:
            return None

def segundos_a_mmss(seg):
    try:
        seg      = int(seg)
        minutos  = seg // 60
        segundos = seg % 60
        return f"{minutos:02d}:{segundos:02d}"
    except Exception:
        return "-"

def extraer_fecha_nombre(nombre_archivo):
    match = re.search(r"(\d{4}-\d{2}-\d{2})", nombre_archivo)
    if match:
        return match.group(1)
    return datetime.today().strftime("%Y-%m-%d")


# ============================================================
# PROCESAMIENTO DE DATOS
# ============================================================
def procesar_datos(df):

    # Solo sub-picklists (#1, #2...)
    mask_subpick   = df[COLUMNA_PICKLIST].astype(str).str.contains("#", na=False)
    df_filtrado    = df[mask_subpick].copy()

    # Solo asociados individuales (sin coma)
    mask_individual = ~df_filtrado[COLUMNA_ASSOCIATE].astype(str).str.contains(",", na=False)
    df_filtrado     = df_filtrado[mask_individual].copy()

    # Eliminar filas sin asociado
    df_filtrado = df_filtrado[df_filtrado[COLUMNA_ASSOCIATE].notna()]
    df_filtrado = df_filtrado[df_filtrado[COLUMNA_ASSOCIATE].astype(str).str.strip() != ""]

    # Conversiones numéricas
    df_filtrado["Duration_seg"] = df_filtrado[COLUMNA_DURATION].apply(duration_a_segundos)
    df_filtrado[COLUMNA_BAGS]   = pd.to_numeric(df_filtrado[COLUMNA_BAGS], errors="coerce").fillna(0)
    df_filtrado[COLUMNA_OVS]    = pd.to_numeric(df_filtrado[COLUMNA_OVS],  errors="coerce").fillna(0)

    # Tabla agrupada
    tabla = df_filtrado.groupby(COLUMNA_ASSOCIATE).agg(
        Num_Picklists    = (COLUMNA_PICKLIST, "count"),
        Total_Bags       = (COLUMNA_BAGS,     "sum"),
        Total_OVs        = (COLUMNA_OVS,      "sum"),
        Avg_Duration_seg = ("Duration_seg",   "mean")
    ).reset_index()

    tabla["Avg_Duration"] = tabla["Avg_Duration_seg"].apply(segundos_a_mmss)
    tabla = tabla.drop(columns=["Avg_Duration_seg"])

    # Marcar rol de cada persona
    def get_rol(login):
        l = str(login).lower()
        if l in [x.lower() for x in MANAGER_LOGINS]:
            return "Manager"
        elif l in [x.lower() for x in YM_LOGINS]:
            return "YM"
        else:
            return "AA"

    tabla["Rol"] = tabla[COLUMNA_ASSOCIATE].apply(get_rol)

    # Media de picklists SOLO con AAs
    tabla_aa  = tabla[tabla["Rol"] == "AA"]
    num_aa    = len(tabla_aa)

    # Picklists totales excluyendo YM y Managers
    total_picklists_aa = tabla_aa["Num_Picklists"].sum()
    media_picklists    = round(total_picklists_aa / num_aa, 1) if num_aa > 0 else 0

    # Columna % sobre la media (solo para AAs)
    def calcular_pct(row):
        if row["Rol"] != "AA":
            return None
        diff = ((row["Num_Picklists"] - media_picklists) / media_picklists) * 100
        return round(diff, 1)

    tabla["Pct_vs_Media"] = tabla.apply(calcular_pct, axis=1)

    # Ordenar: AAs primero por picklists desc, luego YM, luego Managers
    orden_rol = {"AA": 0, "YM": 1, "Manager": 2}
    tabla["_orden"] = tabla["Rol"].map(orden_rol)
    tabla = tabla.sort_values(["_orden", "Num_Picklists"], ascending=[True, False])
    tabla = tabla.drop(columns=["_orden"]).reset_index(drop=True)

    return tabla, media_picklists


# ============================================================
# EXPORT A EXCEL CON FORMATO
# ============================================================
def generar_excel(tabla, media_picklists, fecha):

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        # Preparar datos para escribir
        df_export = tabla.copy()
        df_export["vs Media"] = df_export["Pct_vs_Media"].apply(
            lambda x: f"+{x}%" if (x is not None and x >= 0)
                      else (f"{x}%" if x is not None else "-")
        )
        df_export = df_export.drop(columns=["Pct_vs_Media", "Rol"])
        df_export = df_export.rename(columns={
            COLUMNA_ASSOCIATE: "Asociado",
            "Num_Picklists":   "Picklists",
            "Total_Bags":      "Bags",
            "Total_OVs":       "OVs",
            "Avg_Duration":    "Duración Media"
        })

        # Escribir hoja principal
        df_export.to_excel(writer, sheet_name="Performance", index=False, startrow=1)
        ws = writer.sheets["Performance"]

        # Escribir hoja auxiliar para el gráfico
        tabla_aa = tabla[tabla["Rol"] == "AA"].copy().reset_index(drop=True)
        num_aa_chart = len(tabla_aa)

        datos_grafico = pd.DataFrame({
            "Asociado":  tabla_aa[COLUMNA_ASSOCIATE].values,
            "Picklists": tabla_aa["Num_Picklists"].astype(int).values,
            "Media":     [float(media_picklists)] * num_aa_chart
        })
        datos_grafico.to_excel(writer, sheet_name="_chart_data", index=False)
        ws_data = writer.sheets["_chart_data"]

        # -------------------------------------------------------
        # FORMATO HOJA PRINCIPAL
        # -------------------------------------------------------

        # Título
        ws["A1"] = f"Pick & Stage — AA's Pace   {fecha}"
        ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="1F3864")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A1:{get_column_letter(len(df_export.columns))}1")

        # Cabecera de tabla
        header_fill   = PatternFill("solid", fgColor="2E75B6")
        header_font   = Font(bold=True, color="FFFFFF", size=10)
        header_border = Border(bottom=Side(style="medium", color="FFFFFF"))
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = header_border
        ws.row_dimensions[2].height = 20

        # Colores por rendimiento
        fill_verde_fuerte = PatternFill("solid", fgColor="70AD47")
        fill_verde_suave  = PatternFill("solid", fgColor="C6EFCE")
        fill_amarillo     = PatternFill("solid", fgColor="FFEB9C")
        font_gris         = Font(color="AAAAAA", size=10)
        font_normal       = Font(color="000000", size=10)
        font_verde_fuerte = Font(color="375623", bold=True, size=10)

        thin_border = Border(
            left=Side(style="thin",   color="D0D0D0"),
            right=Side(style="thin",  color="D0D0D0"),
            top=Side(style="thin",    color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0")
        )

        for row_idx, row_data in tabla.iterrows():
            excel_row = row_idx + 3
            rol = row_data["Rol"]
            pct = row_data["Pct_vs_Media"]

            if rol in ["YM", "Manager"]:
                fila_fill = PatternFill("solid", fgColor="F2F2F2")
                fila_font = font_gris
            elif pct is None:
                fila_fill = PatternFill("solid", fgColor="F2F2F2")
                fila_font = font_gris
            elif pct > 10:
                fila_fill = fill_verde_fuerte
                fila_font = font_verde_fuerte
            elif pct >= -10:
                fila_fill = fill_verde_suave
                fila_font = font_normal
            else:
                fila_fill = fill_amarillo
                fila_font = font_normal

            for col_idx in range(1, len(df_export.columns) + 1):
                cell           = ws.cell(row=excel_row, column=col_idx)
                cell.fill      = fila_fill
                cell.font      = fila_font
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[excel_row].height = 18

        # Colorear columna vs Media
        col_vs = list(df_export.columns).index("vs Media") + 1
        for row_idx, row_data in tabla.iterrows():
            excel_row = row_idx + 3
            pct = row_data["Pct_vs_Media"]
            cell = ws.cell(row=excel_row, column=col_vs)
            if pct is not None:
                if pct >= 0:
                    cell.font = Font(color="375623", bold=True, size=10)
                else:
                    cell.font = Font(color="9C5700", bold=True, size=10)

        # Ajustar ancho de columnas
        for col_idx in range(1, len(df_export.columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_ancho = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(tabla) + 3)
            )
            ws.column_dimensions[col_letter].width = max(max_ancho + 4, 12)

        # Leyenda
        fila_leyenda = len(tabla) + 4
        ws.cell(row=fila_leyenda,     column=1).value = "Leyenda:"
        ws.cell(row=fila_leyenda,     column=1).font  = Font(bold=True, size=9)
        ws.cell(row=fila_leyenda + 1, column=1).value = "Verde oscuro: >10% sobre la media"
        ws.cell(row=fila_leyenda + 1, column=1).fill  = fill_verde_fuerte
        ws.cell(row=fila_leyenda + 1, column=1).font  = Font(size=9)
        ws.cell(row=fila_leyenda + 2, column=1).value = "Verde claro: dentro de la media (±10%)"
        ws.cell(row=fila_leyenda + 2, column=1).fill  = fill_verde_suave
        ws.cell(row=fila_leyenda + 2, column=1).font  = Font(size=9)
        ws.cell(row=fila_leyenda + 3, column=1).value = "Amarillo: >10% por debajo de la media"
        ws.cell(row=fila_leyenda + 3, column=1).fill  = fill_amarillo
        ws.cell(row=fila_leyenda + 3, column=1).font  = Font(size=9)
        ws.cell(row=fila_leyenda + 4, column=1).value = "Gris: YM o Manager (no se evalúa)"
        ws.cell(row=fila_leyenda + 4, column=1).fill  = PatternFill("solid", fgColor="F2F2F2")
        ws.cell(row=fila_leyenda + 4, column=1).font  = Font(color="AAAAAA", size=9)
        ws.cell(row=fila_leyenda + 5, column=1).value = f"Media AA: {media_picklists} picklists"
        ws.cell(row=fila_leyenda + 5, column=1).font  = Font(bold=True, size=9, color="C55A11")

        # -------------------------------------------------------
        # GRÁFICO — referencias desde hoja _chart_data
        # -------------------------------------------------------
        from openpyxl.chart import LineChart

        # Serie de barras: columna B (Picklists), filas 1 a num_aa_chart+1
        data_bars = Reference(
            ws_data,
            min_col=2, max_col=2,
            min_row=1, max_row=num_aa_chart + 1
        )
        # Categorías: columna A (Asociado), filas 2 a num_aa_chart+1
        cats = Reference(
            ws_data,
            min_col=1, max_col=1,
            min_row=2, max_row=num_aa_chart + 1
        )
        # Serie de línea: columna C (Media), filas 1 a num_aa_chart+1
        data_line = Reference(
            ws_data,
            min_col=3, max_col=3,
            min_row=1, max_row=num_aa_chart + 1
        )

        # Gráfico de barras
        chart = BarChart()
        chart.type      = "col"
        chart.grouping  = "clustered"
        chart.title     = f"Picklists por AA  |  Media: {media_picklists}"
        chart.y_axis.title = "Picklists"
        chart.x_axis.title = "Asociado"
        chart.width     = 26
        chart.height    = 19        # Más alto para que quepan los nombres abajo
        chart.style     = 10
        chart.add_data(data_bars, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill      = "2E75B6"
        chart.series[0].graphicalProperties.line.solidFill = "2E75B6"

        # Etiquetas de valor encima de cada barra
        from openpyxl.chart.label import DataLabelList
        chart.series[0].dLbls = DataLabelList()
        chart.series[0].dLbls.showVal   = True
        chart.series[0].dLbls.showSerName = False
        chart.series[0].dLbls.showCatName = False
        chart.series[0].dLbls.showLegendKey = False

        # Rotar etiquetas del eje X para que se lean los nombres
        chart.x_axis.tickLblPos = "low"
        from openpyxl.chart.axis import NumericAxis
        chart.x_axis.txPr = None
        chart.plot_area.layout = None

        # Aplicar rotación de -45 grados a las etiquetas del eje X
        from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
        from lxml import etree
        tx_pr = etree.fromstring(
            '<txPr xmlns="http://schemas.openxmlformats.org/drawingml/2006/chart">'
            '<a:bodyPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" rot="-2700000" vert="horz"/>'
            '<a:lstStyle xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"/>'
            '<a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            '<a:pPr>'
            '<a:defRPr sz="700" b="0">'
            '<a:solidFill><a:srgbClr val="000000"/></a:solidFill>'
            '</a:defRPr>'
            '</a:pPr>'
            '</a:p>'
            '</txPr>'
        )
        chart.x_axis._txPr = tx_pr
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.delete = False
        chart.x_axis.noMultiLvlLbl = True



        # Línea naranja de media
        line_chart = LineChart()
        line_chart.add_data(data_line, titles_from_data=True)
        line_chart.set_categories(cats)
        line_chart.series[0].graphicalProperties.line.solidFill = "FF6600"
        line_chart.series[0].graphicalProperties.line.width     = 30000
        line_chart.series[0].smooth = False

        chart += line_chart

        fila_inicio_chart = fila_leyenda + 7
        ws.add_chart(chart, f"A{fila_inicio_chart}")

        # Ocultar hoja auxiliar
        ws_data.sheet_state = "hidden"

    buffer.seek(0)
    return buffer



# ============================================================
# GRÁFICOS PLOTLY PARA STREAMLIT
# ============================================================
def grafico_plotly(tabla, columna, titulo, color_barra, media=None):
    df_plot = tabla[tabla["Rol"] == "AA"].sort_values(columna, ascending=False)

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=df_plot[COLUMNA_ASSOCIATE],
        y=df_plot[columna],
        marker_color=color_barra,
        name=titulo
    ))

    if media is not None:
        fig.add_hline(
            y=media,
            line_color="orange",
            line_width=2.5,
            line_dash="dash",
            annotation_text=f"Media: {media}",
            annotation_position="top right",
            annotation_font_color="orange"
        )

    fig.update_layout(
        xaxis_title="Asociado",
        yaxis_title=titulo,
        plot_bgcolor="white",
        xaxis=dict(tickangle=-35),
        margin=dict(t=30, b=60),
        height=380
    )

    return fig


# ============================================================
# INTERFAZ STREAMLIT
# ============================================================
st.set_page_config(
    page_title="Pick & Stage — AA's Pace DCZ4",
    page_icon="📦",
    layout="wide"
)

st.title("📦 Pick & Stage — AA's Pace")
st.markdown("Sube el archivo PICK del día para obtener el análisis de rendimiento por asociado.")
st.divider()

archivo = st.file_uploader(
    "Selecciona el archivo PICK (.xlsx o .csv)",
    type=["xlsx", "xls", "csv"]
)

if archivo is not None:
    nombre    = archivo.name
    extension = os.path.splitext(nombre)[1].lower()
    fecha     = extraer_fecha_nombre(nombre)

    try:
        with st.spinner("Procesando datos..."):
            if extension in [".xlsx", ".xls"]:
                df = pd.read_excel(archivo, sheet_name=0, header=1)
            else:
                df = pd.read_csv(archivo)

            tabla, media_picklists = procesar_datos(df)

        tabla_aa = tabla[tabla["Rol"] == "AA"]

        st.success(f"✅ Procesado correctamente — {len(tabla_aa)} AAs encontrados | Media: **{media_picklists} picklists**")
        st.divider()

        # --- Métricas resumen
        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("AAs activos",      len(tabla_aa))
        col2.metric("Media Picklists",  media_picklists)
        col3.metric("Total Bags",       int(tabla_aa["Total_Bags"].sum()))
        col4.metric("Total OVs",        int(tabla_aa["Total_OVs"].sum()))
        col5.metric("Total Picklists",  int(tabla_aa["Num_Picklists"].sum()))

        st.divider()

        # --- Tabla con colores en Streamlit
        st.subheader("📊 Rendimiento por Asociado")

        def colorear_fila(row):
            # Parseamos el valor desde "vs Media" en lugar de "Pct_vs_Media"
            try:
                pct = float(str(row["vs Media"]).replace("%", "").replace("+", ""))
            except Exception:
                pct = None

            rol = row.get("Rol", "AA")

            if rol in ["YM", "Manager"]:
                return ["color: grey"] * len(row)
            elif pct is None:
                return [""] * len(row)
            elif pct > 10:
                return ["background-color: #70AD47; color: #375623; font-weight: bold"] * len(row)
            elif pct >= -10:
                return ["background-color: #C6EFCE"] * len(row)
            else:
                return ["background-color: #FFEB9C"] * len(row)


        tabla_display = tabla.copy()
        tabla_display["vs Media"] = tabla_display["Pct_vs_Media"].apply(
            lambda x: f"+{x}%" if (x is not None and x >= 0)
                      else (f"{x}%" if x is not None else "-")
        )
        tabla_display = tabla_display.drop(columns=["Pct_vs_Media"])

        st.dataframe(
            tabla_display.style.apply(colorear_fila, axis=1),
            use_container_width=True,
            hide_index=True
        )

        st.divider()

        # --- Gráficos ordenados
        st.subheader("📈 Visualizaciones (solo AAs, orden: mayor a menor)")

        tab1, tab2, tab3 = st.tabs(["Picklists", "Bags", "OVs"])

        with tab1:
            st.plotly_chart(
                grafico_plotly(tabla, "Num_Picklists", "Picklists", "#2E75B6", media=media_picklists),
                use_container_width=True
            )
        with tab2:
            st.plotly_chart(
                grafico_plotly(tabla, "Total_Bags", "Bags", "#21C55D"),
                use_container_width=True
            )
        with tab3:
            st.plotly_chart(
                grafico_plotly(tabla, "Total_OVs", "OVs", "#F97316"),
                use_container_width=True
            )

        st.divider()

        # --- Descarga Excel
        excel_buffer = generar_excel(tabla, media_picklists, fecha)
        nombre_salida = f"Pick&Stage AA's pace {fecha}.xlsx"

        st.download_button(
            label="⬇️ Descargar Excel completo",
            data=excel_buffer,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")
        st.exception(e)

else:
    st.info("👆 Sube un archivo PICK para comenzar.")
